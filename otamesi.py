from time import sleep,time
from selenium import webdriver 
import requests 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
import ffmpeg
import os
import openpyxl
import pprint
import datetime
import urllib.request, urllib.error
from bs4 import BeautifulSoup
import socks, socket
import difflib





# socks.set_default_proxy(socks.PROXY_TYPE_SOCKS5, '127.0.0.1', 9050)
# socket.socket = socks.socksocket

# # URLからHTMLを返す
# def fetch_html(url):
#   res = urllib.request.urlopen(url)
#   return BeautifulSoup(res, 'html.parser')

# # 現在のグローバルIPアドレスを返す
# def get_ip_addr():
#   html = fetch_html('http://checkip.dyndns.com/')
#   return html.body.text.split(': ')[1]

# # Torを使っているかを返す
# def check_use_tor():
#   html = fetch_html('https://check.torproject.org/')
#   return html.find('h1')['class'][0] != 'off'

# print('You are using tor.' if check_use_tor() else 'You are not using tor.')
# print('Current IP address is ' + get_ip_addr())


# dt_now = datetime.datetime.now()
# print(dt_now.strftime('%Y年%m月%d日 %H:%M:%S'))
# dt2 = datetime.timedelta(minutes=10)

# dt_10leter = dt_now + dt2
# print(dt_10leter.strftime('%H:%M %p'))
# print(dt_now.strftime('%H:%M %p'))

# print(dt_now.month)
# print(dt_now.year)
# # this_month = str(dt_now.month) + "月"
# this_month = int(dt_now.month) - 1 
# print(this_month)

# today = dt_now.day 
# print(today)


def time():

    for i in range(10):
        print(i)
        dt_now = datetime.datetime.now()
        # print(dt_now.strftime('%Y年%m月%d日 %H:%M:%S'))
        dt2 = datetime.timedelta(minutes= 10 + (50000 * i) )

        dt_10leter = dt_now + dt2
        this_month = int(dt_now.month) -1
        # print(dt_10leter.strftime('%Y/%D %H:%M %p'))
        print(dt_10leter.strftime('%Y年%m月%d日 %H:%M:%S'))
        print(dt_10leter.year)
        
        # print(dt_now.strftime('%H:%M %p'))

        # print(dt_now.month)
        # print(dt_now.year)
        # # this_month = str(dt_now.month) + "月"
        # this_month = int(dt_now.month) - 1 
        # print(this_month)

        # today = dt_now.day 
        # print(today)

time()



wb = openpyxl.load_workbook('/Users/yoshizawamasaaki/Documents/sample_movie/MgsSampleMp4.xltx')

ws = wb.worksheets[0]

afe_links = []
for cell in ws['A']:
    afe_links.append(cell.value)
# print(afe_links)

for cell in ws['B']:
    afe_links.append(cell.value)
# print(afe_links)



movie_info_list = {}
for k, v in zip(ws['A'],ws['B']):
    movie_info_list[k.value] = v.value


# movie_title = "結衣.mp4"
# if  ".mp4" in movie_title:
#     movie_title = movie_title.replace('.mp4','')
#     print(movie_title)

# for key, value in movie_info_list.items():
#     if key == movie_title:
#         print(value)
#         print('成功')




    # movie_info_list.append(row_dic)

# pprint.pprint(movie_info_list)
# pprint.pprint(row_dic)

# print(row_dic["https://www.mgstage.com/product/product_detail/435MFC-030/?agef=1&utm_medium=mgs_affiliate&utm_source=mgs_affiliate_linktool&aff=TXJHY5AUPQWYLHZSB5FDHOR6P4&utm_campaign=mgs_affiliate_linktool&utm_content=TXJHY5AUPQWYLHZSB5FDHOR6P4&form=mgs_asp_linktool_TXJHY5AUPQWYLHZSB5FDHOR6P4"])
# print(row_dic.get['https://www.mgstage.com/product/product_detail/435MFC-043/?agef=1&utm_medium=mgs_affiliate&utm_source=mgs_affiliate_linktool&aff=TXJHY5AUPQWYLHZSB5FDHOR6P4&utm_campaign=mgs_affiliate_linktool&utm_content=TXJHY5AUPQWYLHZSB5FDHOR6P4&form=mgs_asp_linktool_TXJHY5AUPQWYLHZSB5FDHOR6P4'])
# print(row_dic["雫"])

# mydict = {"JP":"Japan", "DE":"Germany", "FR":"France"}




i = 0
afe_link_dic = {}
for afe_link in afe_links:
    afe_link_dic[i] = afe_link 
    i += 1


# 動画リスト

# 列名のセル
header_cells = None

# for row in ws.rows:
#     if row[0].row == 1:
#         # １行目
#         header_cells = row
#     else:
#         # ２行目以降
#         row_dic = {}
#         # セルの値を「key-value」で登録
#         for k, v in zip(header_cells, row):
#             row_dic[k.value] = v.value
#         movie_info_list.append(row_dic)

# pprint.pprint(movie_info_list)

# for key, value in movie_info_list.items():
#     print(key)
#     print(value)


   

# for j in range(10):
    # pprint.pprint(afe_link_dic)
    # print(afe_link_dic)

def movie_title_conf_1():
    movie_title = "【しろうとハメ撮り】色気と性欲が溢れ出るファビュラスな人妻と不倫ハメ撮り！※のぞみ35歳人妻.mp4"
    if  ".mp4" in movie_title:
                    movie_title = movie_title.replace('.mp4','')
                    print(movie_title)

    for key, value in afe_link_dic.items():
        # pprint.pprint(value)
        # print(str(key)+".mp4")
        # mozi = str(key)+".mp4"
        mozi = value
        # print(mozi)
        if mozi == movie_title:
            print(mozi)
            print(afe_link_dic[key])


def movie_title_conf2():
    # movie_title = "ゆあぁず"
    movie_title = "ゆあぁず"
   
# ゆあぁず

# ゆあぁず
str1 = "ゆあぁず"
str2 = "ゆあぁず"

s = difflib.SequenceMatcher(None, str1, str2).ratio()

print (str1, "<~>", str2)
print ("match ratio:", s, "\n")



if "ゆあぁず" == "ゆあぁず":
    print(true)


    for key, value in movie_info_list.items():
        if key == movie_title:
            print('タイトル合致成功')
            print(value)

    

movie_title_conf2()



def search():

    options = webdriver.ChromeOptions()


    # 2. シークレットモードでの使用
    options.add_argument('--incognito')

    # 3. User-Agentの設定
    options.add_argument('--user-agent= Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36')



    #step1 : driverを作成する
    driver = webdriver.Chrome(
        executable_path='/Users/yoshizawamasaaki/Desktop/lessonn/scraping/tools/chromedriver',
        options=options
    )
    driver.implicitly_wait(10)

    dt_now = datetime.datetime.now()
    dt2 = datetime.timedelta(minutes=10)
    dt_10leter = dt_now + dt2 
    this_year = dt_now.year

    driver.get('https://news.yahoo.co.jp')
    sleep(2)

    search_box = driver.find_element_by_css_selector('input.sc-kgoBCf') 

  
    search_box.send_keys(dt_10leter.strftime('%H:%M %p'))
    search_box.send_keys(dt_10leter.strftime('%H:%M %p'))
    sleep(2)

    search_box.send_keys(Keys.COMMAND, "a")
    search_box.send_keys(Keys.DELETE)
    

    # search_box.send_keys('機械学習')
    sleep(2)

    search_box.submit()
    sleep(2)

# search()


# def scroll_down():
#     """A method for scrolling the page."""
#     # Get scroll height.
#     last_height= driver.execute_script("return document.body.scrollHeight")
#     while True:
#         # Scroll down to the bottom.
        
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         # Wait to load the page.
#         button_tag = driver.find_element_by_css_selector('span.button-simple > button')
#         button_tag.click() 
#         sleep(2)
#         # Calculate new scroll height and compare with last scroll height.
#         new_height= driver.execute_script("return document.body.scrollHeight")
#         print("new_height:",new_height)
#         print("last_height:",last_height)
#         if new_height== last_height:
#             break
#         last_height= new_height
        

# scroll_down()

# print('成功しました')




# movie_download_name = "【しろうとハメ撮り】美乳巨乳でわがままボディなエロ娘とハメ撮り/和佳/23歳/Gカップ(巨乳).mp4"

# if  "/" in movie_download_name :
#     movie_download_name = movie_download_name.replace('/','')
#     print(movie_download_name)
